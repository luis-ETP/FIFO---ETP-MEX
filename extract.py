"""
Extract structured data from a processed FIFO workbook for the dashboard.
v2 - includes investment summary extraction
"""
import openpyxl
from collections import defaultdict

LITERS_PER_GAL = 3.7854

def extract(path, src_path=None):
    wb     = openpyxl.load_workbook(path, data_only=True)
    wb_src = openpyxl.load_workbook(src_path, data_only=True) if src_path else wb

    overall_summary = _extract_overall_summary(wb)
    inventory       = _extract_inventory(wb)
    fifo_rows       = _extract_fifo(wb)
    meta            = _extract_meta(wb)

    bol_tab    = _extract_bol(wb_src)
    investment = _extract_investment_summary(wb_src)
    return overall_summary, inventory, fifo_rows, meta, investment, bol_tab

# ── Overall Summary ────────────────────────────────────────────────────────────
def _extract_overall_summary(wb):
    ws = wb["Overall Summary"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains "Row Labels")
    hdr_idx = next(i for i, r in enumerate(rows) if r[0] == "Row Labels")
    headers = [str(v).strip() if v else f"col{j}" for j, v in enumerate(rows[hdr_idx])]

    result = []
    for row in rows[hdr_idx + 1:]:
        if not row[0]: continue
        entry = {}
        for j, h in enumerate(headers):
            v = row[j]
            entry[h] = float(v) if isinstance(v, (int, float)) else (str(v) if v else None)
        # Normalise key names so dashboard always finds them regardless of Excel naming
        entry["_wired"]     = entry.get("Total Wired Amount", 0) or 0
        entry["_paid_gal"]  = entry.get("Paid for Gallons (Allocation)") or entry.get("Paid for Gallons ") or 0
        entry["_pulled"]    = entry.get("Gallons Pulled from Allocation (RTB & RTC)") or entry.get("Gallons Pulled (RTB & RTC)") or 0
        entry["_rem_alloc"] = entry.get("Remaining Gallons in Allocation") or 0
        entry["_rem_inv"]   = entry.get("Remaining Gallons in Inventory") or 0
        entry["_avg_cost"]  = entry.get("Weighted Average Cost in Inventory (MXN/L)") or entry.get("Weighted Average Cost in Inventory") or 0
        entry["_paid_back"] = entry.get("Amount Paid Back by Mexico (MXN)") or entry.get("Amount Paid Back by Mexico ") or entry.get("Amount Paid Back by Mexico") or 0
        entry["_balance"]   = entry.get("Mexico Balance (MXN)") or entry.get("Mexico Balance") or 0
        result.append(entry)
    return result

# ── Inventory (from FIFO sheet + Purchase to BOL-RTB) ─────────────────────────
def _extract_inventory(wb):
    """
    Build hierarchy: bulk_plant -> product -> batch -> supplier -> invoice -> [bols]

    BOLs that span two batches/invoices (e.g. Batch="1 | 2", Invoice="A | B")
    are split into separate entries — one per batch/invoice — with proportional liters.
    Each BOL always appears INDIVIDUALLY under its own invoice.
    """
    # Step 1: read FIFO sheet for each RTB BOL
    ws_fifo = wb["FIFO"]
    fifo_headers = [str(v).strip() if v else f"col{j}"
                    for j, v in enumerate(next(ws_fifo.iter_rows(values_only=True)))]

    fifo_bols = {}  # bol_str -> {liters, remaining_l, cost_per_l, bp, prod}
    for row in ws_fifo.iter_rows(min_row=2, values_only=True):
        if not row[0]: break
        entry = {h: row[j] for j, h in enumerate(fifo_headers)}
        if entry.get("Type") != "RTB": continue
        bol = str(entry.get("BOL", "") or "")
        if not bol: continue
        liters = float(entry.get("Liters", 0) or 0)
        rem    = entry.get("Remaining L (BOL)")
        fifo_bols[bol] = {
            "liters":      liters,
            "remaining_l": float(rem) if rem is not None else liters,
            "cost_per_l":  float(entry.get("Cost / L (MXN)", 0) or 0),
            "bp":          str(entry.get("Bulk Plant", "") or ""),
            "prod":        str(entry.get("Product", "") or ""),
        }

    # Step 2: read Purchase to BOL-RTB for invoice, batch, supplier per BOL
    # For split BOLs (e.g. Batch="1 | 2", Invoice="A | B"), split into two entries
    bol_entries = []  # list of {bol, batch, invoice, supplier, alloc_frac}
    try:
        ws_bol = wb["Purchase to BOL-RTB"]
        for row in ws_bol.iter_rows(min_row=8, values_only=True):
            if not row[2]: break
            bol_str  = str(row[5]).strip() if row[5] else ""
            supplier = str(row[2]).strip() if row[2] else ""
            inv_raw  = str(row[3]).strip() if row[3] else ""
            bat_raw  = str(row[4]).strip() if row[4] else ""
            cost_j   = float(row[9])  if row[9]  is not None else 0.0  # J Cost/Gal USD
            cost_l_raw = str(row[11]).strip() if row[11] is not None else ""

            if not bol_str: continue

            # Split multi-batch/invoice BOLs into individual entries
            batches  = [b.strip() for b in bat_raw.split("|")] if "|" in bat_raw else [bat_raw]
            invoices = [v.strip() for v in inv_raw.split("|")] if "|" in inv_raw else [inv_raw]

            if len(batches) > 1 or len(invoices) > 1:
                # Cross-batch BOL: split proportionally
                # We don't have exact split fractions here, so split equally
                # (the FIFO sheet has the blended cost already)
                n = max(len(batches), len(invoices))
                for k in range(n):
                    b = batches[k] if k < len(batches) else batches[-1]
                    inv = invoices[k] if k < len(invoices) else invoices[-1]
                    bol_entries.append({
                        "bol": bol_str, "batch": b, "invoice": inv,
                        "supplier": supplier, "split_n": n, "split_k": k,
                    })
            else:
                bol_entries.append({
                    "bol": bol_str, "batch": bat_raw, "invoice": inv_raw,
                    "supplier": supplier, "split_n": 1, "split_k": 0,
                })
    except Exception as e:
        pass

    # Step 3: build hierarchy
    result = {}
    for entry in bol_entries:
        bol      = entry["bol"]
        batch    = entry["batch"]
        invoice  = entry["invoice"]
        supplier = entry["supplier"]
        split_n  = entry["split_n"]
        fifo     = fifo_bols.get(bol)
        if not fifo: continue

        bp   = fifo["bp"]
        prod = fifo["prod"]
        # Proportional liters for split BOLs
        liters      = round(fifo["liters"]      / split_n, 4)
        remaining_l = round(fifo["remaining_l"] / split_n, 4)
        cost_per_l  = fifo["cost_per_l"]  # blended cost same for all splits

        # Navigate: bp -> prod -> batch -> supplier -> invoice -> [bols]
        result.setdefault(bp, {})
        result[bp].setdefault(prod, {})
        result[bp][prod].setdefault(batch, {})
        result[bp][prod][batch].setdefault(supplier, {})
        result[bp][prod][batch][supplier].setdefault(invoice, [])
        result[bp][prod][batch][supplier][invoice].append({
            "bol":         bol,
            "liters":      liters,
            "remaining_l": remaining_l,
            "cost_per_l":  cost_per_l,
        })

    return result

# ── FIFO rows ──────────────────────────────────────────────────────────────────
def _extract_fifo(wb):
    ws = wb["FIFO"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v else f"col{j}" for j, v in enumerate(rows[0])]

    result = []
    for row in rows[1:]:
        if not row[0]: break
        entry = {}
        for j, h in enumerate(headers):
            v = row[j]
            if hasattr(v, 'isoformat'):
                v = v.strftime("%d/%m/%Y")
            elif isinstance(v, float):
                v = round(v, 4)
            entry[h] = v
        result.append(entry)
    return result

# ── Meta ───────────────────────────────────────────────────────────────────────
def _extract_meta(wb):
    # Pull key KPIs from Overall Summary total row
    ws = wb["Overall Summary"]
    total_row = None
    for row in ws.iter_rows(values_only=True):
        if row[0] and "TOTAL" in str(row[0]).upper():
            total_row = row
            break

    meta = {}
    if total_row:
        def _f(v): 
            try: return round(float(v), 4)
            except: return 0
        # Use overall_summary normalised keys for reliability
        os_rows = _extract_overall_summary(wb)
        total_os = next((r for r in os_rows if r.get("Row Labels","").upper().find("TOTAL") >= 0), {})
        meta = {
            "total_invoiced_usd":      _f(total_row[1]),
            "total_gallons":           _f(total_row[2]),
            "total_wired":             total_os.get("_wired", _f(total_row[3])),
            "paid_for_gallons":        total_os.get("_paid_gal", _f(total_row[4])),
            "gallons_pulled":          total_os.get("_pulled", _f(total_row[5])),
            "remaining_allocation":    total_os.get("_rem_alloc", _f(total_row[6])),
            "remaining_inventory_gal": total_os.get("_rem_inv", _f(total_row[7])),
            "avg_cost_inventory":      total_os.get("_avg_cost", _f(total_row[8])),
            "amount_paid_back":        total_os.get("_paid_back", _f(total_row[9])),
            "mexico_balance":          total_os.get("_balance", _f(total_row[10])),
        }
    return meta

# ── Investment Summary ─────────────────────────────────────────────────────────
def _extract_investment_summary(wb, uploaded_at=None):
    """
    Read Investment Summary values directly from cached Excel formula results.
    Falls back to computing from raw columns when cache is missing.
    """
    f  = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
    fz = lambda v: float(v) if isinstance(v, (int, float)) else None

    ws = wb["Investment Summary"]
    rows = {i: row for i, row in enumerate(ws.iter_rows(values_only=True), start=1)}

    # ── Date (G3) ────────────────────────────────────────────────────────────
    r3 = rows.get(3, [])
    as_of_val = r3[6] if len(r3) > 6 else None
    if hasattr(as_of_val, "strftime"):
        as_of_str = as_of_val.strftime("%d-%b-%Y")
    elif as_of_val:
        as_of_str = str(as_of_val)
    else:
        as_of_str = uploaded_at or ""

    # ── Committed Capital (rows 7-8) ─────────────────────────────────────────
    commits = []
    for ri in (7, 8):
        row = rows.get(ri, [])
        if len(row) < 6: continue
        usd = f(row[2]); mxn = f(row[3]); fx = f(row[5])
        dv  = row[4]
        dstr = dv.strftime("%d-%b-%y") if hasattr(dv, "strftime") else str(dv or "")
        commits.append({"round": str(row[1] or ""), "usd": usd, "mxn": mxn, "date": dstr, "fx": fx})

    r10 = rows.get(10, [])
    total_committed_usd = f(r10[2]) if len(r10) > 2 else sum(c["usd"] for c in commits)
    total_committed_mxn = f(r10[3]) if len(r10) > 3 else sum(c["mxn"] for c in commits)
    avg_fx = total_committed_mxn / total_committed_usd if total_committed_usd else 17.31

    # ── KPI row 14 ───────────────────────────────────────────────────────────
    r14 = rows.get(14, [])
    active_capital  = f(r14[2]) if len(r14) > 2 else 0.0
    available       = f(r14[3]) if len(r14) > 3 else 0.0
    recovered_mxn   = f(r14[4]) if len(r14) > 4 else 0.0
    revolved        = f(r14[5]) if len(r14) > 5 else 0.0
    total_margin    = f(r14[6]) if len(r14) > 6 else 0.0
    investor_share  = f(r14[7]) if len(r14) > 7 else total_margin * 0.40
    inv_share_pct   = investor_share / total_margin if total_margin else 0.40

    # ── Active Capital detail (rows 18-23) ───────────────────────────────────
    def _rd(ri, ci, cj=None):
        r = rows.get(ri, [])
        v1 = f(r[ci]) if len(r) > ci else 0.0
        v2 = f(r[cj]) if cj is not None and len(r) > cj else 0.0
        return v1, v2

    alloc_mxn,    alloc_liters    = _rd(18, 2, 3)
    inv_mxn,      inv_liters      = _rd(19, 2, 3)
    rtc_pend_mxn, rtc_pend_liters = _rd(20, 2, 3)
    btc_pend_mxn, btc_pend_liters = _rd(21, 2, 3)
    total_act_mxn, total_act_lit  = _rd(23, 2, 3)

    # ── Recovered Capital detail (rows 27-30) ────────────────────────────────
    def _rec(ri):
        r = rows.get(ri, [])
        return {
            "tc":     f(r[2]) if len(r) > 2 else 0.0,
            "liters": f(r[3]) if len(r) > 3 else 0.0,
            "margin": f(r[4]) if len(r) > 4 else 0.0,
            "roi":    f(r[5]) if len(r) > 5 else 0.0,
            "mxnl":   f(r[6]) if len(r) > 6 else 0.0,
            "usdgal": f(r[7]) if len(r) > 7 else 0.0,
        }

    rtc_rec   = _rec(27)
    btc_rec   = _rec(28)
    total_rec = _rec(30)

    return {
        "as_of":               as_of_str,
        "commits":             commits,
        "total_committed_usd": total_committed_usd,
        "total_committed_mxn": total_committed_mxn,
        "active_capital":      active_capital,
        "available":           available,
        "recovered_mxn":       recovered_mxn,
        "revolved":            revolved,
        "total_margin":        total_margin,
        "investor_share":      investor_share,
        "inv_share_pct":       inv_share_pct,
        "active_detail": {
            "alloc_mxn":       alloc_mxn,    "alloc_liters":    alloc_liters,
            "inv_mxn":         inv_mxn,      "inv_liters":      inv_liters,
            "rtc_pend_mxn":    rtc_pend_mxn, "rtc_pend_liters": rtc_pend_liters,
            "btc_pend_mxn":    btc_pend_mxn, "btc_pend_liters": btc_pend_liters,
            "total_mxn":       total_act_mxn,"total_liters":    total_act_lit,
        },
        "recovered_detail": {
            "rtc":   rtc_rec,
            "btc":   btc_rec,
            "total": total_rec,
        },
    }

# ── Purchase to BOL-RTB ────────────────────────────────────────────────────────
def _extract_bol(wb):
    ws = wb["Purchase to BOL-RTB"]
    f  = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
    fs = lambda v: str(v).strip() if v is not None else ""

    # Summary from top-right corner (rows 2-5, col R = index 17)
    rows = list(ws.iter_rows(values_only=True))
    summary = {
        "total_invoiced":      f(rows[1][17]),   # row 2 col R
        "received_payments":   f(rows[2][17]),   # row 3
        "open_balance":        f(rows[3][17]),   # row 4
        "total_not_invoiced":  f(rows[4][17]),   # row 5
    }

    # Data rows (row 8+)
    bols = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[2]: break
        inv  = fs(row[17])   # R Invoice#
        col4 = f(row[22])    # Column4: 1 if invoiced
        col5 = f(row[23])    # Column5: 1 if balance=0
        col6 = f(row[24])    # Column6: col4+col5

        # Balance status: 'paid' | 'open' | 'not_invoiced'
        if col4 == 0:
            status = "not_invoiced"
        elif col5 == 1:
            status = "paid"
        else:
            status = "open"

        bols.append({
            "bol":         fs(row[5]),
            "gallons":     round(f(row[6]), 2),
            "liters":      round(f(row[7]), 2),
            "product":     fs(row[8]),
            "cost_gal":    round(f(row[16]), 4),   # Tota Cost/Gal (USD)
            "invoice":     inv,
            "inv_amount":  round(f(row[18]), 2),
            "received":    round(f(row[20]), 2),
            "balance":     round(f(row[21]), 2),
            "status":      status,
        })

    return {"summary": summary, "rows": bols}
