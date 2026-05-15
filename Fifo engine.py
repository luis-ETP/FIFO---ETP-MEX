import shutil
from collections import deque, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
 
def run_fifo(SRC, DST):
    SRC = "/mnt/user-data/uploads/Investor_Summary_FIFO_7.xlsx"
    DST = "/mnt/user-data/outputs/Investor_Summary_FIFO.xlsx"
 
    shutil.copy(SRC, DST)
    wb_r = load_workbook(SRC, data_only=True)
    wb   = load_workbook(DST)
 
    ws_inv_r = wb_r["Supplier Invoices"]
    ws_bol_r = wb_r["Purchase to BOL-RTB"]
    ws_lt_r  = wb_r["Load Tracking"]
    ws_inv   = wb["Supplier Invoices"]
    ws_bol   = wb["Purchase to BOL-RTB"]
    ws_lt    = wb["Load Tracking"]
 
    def weighted_avg(pairs):
        total = sum(g for g, _ in pairs)
        return sum(g * r for g, r in pairs) / total if total else 0.0
 
    def join_unique(items):
        seen, out = set(), []
        for x in items:
            s = str(x).strip()
            if s and s not in seen:
                seen.add(s); out.append(s)
        return " | ".join(out)
 
    # ══════════════════════════════════════════════════════════════════════════════
    # Build supplier invoice FIFO queues
    # Supplier Invoices (0-indexed): A(0)=Batch, C(2)=Supplier, D(3)=Invoice#,
    #   G(6)=PaidForGallons, L(11)=Rate(usd/gal), N(13)=SupplyCostDashFuel(MXN/L)
    # ══════════════════════════════════════════════════════════════════════════════
    supplier_queues = defaultdict(list)   # supplier_upper → [entry, ...]
    inv_entries     = {}                  # inv_num → entry
 
    for i, row in enumerate(ws_inv_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if row[0] is None: break
        supplier        = str(row[2]).strip().upper() if row[2] else ""
        inv_num         = str(row[3]).strip()         if row[3] else ""
        rate_usd_gal    = float(row[11])              if row[11] else 0.0
        supply_cost_mxn = float(row[13])              if row[13] else 0.0
        batch           = row[0]
        try:
            paid_gals = float(row[6])
        except (TypeError, ValueError):
            paid_gals = 0.0
        if paid_gals <= 0:
            continue
        entry = {
            "inv_num":         inv_num,
            "batch":           str(batch),
            "rate_usd_gal":    rate_usd_gal,
            "supply_cost_mxn": supply_cost_mxn,
            "avail":           paid_gals,
            "orig":            paid_gals,
            "drawn":           0.0,
            "excel_row":       i,
        }
        supplier_queues[supplier].append(entry)
        inv_entries[inv_num] = entry
 
    # ══════════════════════════════════════════════════════════════════════════════
    # Read Purchase to BOL-RTB into a lookup by BOL number
    # C(2)=Supplier, E(4)=BOL, I(8)=Gallons  (0-indexed)
    # ══════════════════════════════════════════════════════════════════════════════
    bol_info = {}   # bol_str → {supplier, gals, excel_row}
    for i, row in enumerate(ws_bol_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[0]: break
        bol_str  = str(row[4]).strip() if row[4] else ""
        supplier = str(row[2]).strip().upper() if row[2] else ""
        try:
            gals = float(row[8]) if row[8] else 0.0
        except (TypeError, ValueError):
            gals = 0.0
        if bol_str:
            bol_info[bol_str] = {"supplier": supplier, "gals": gals, "excel_row": i}
 
    # ══════════════════════════════════════════════════════════════════════════════
    # STAGE 1 — Allocate supplier invoices in LOAD TRACKING sheet order
    # This is the master order for FIFO: RTB/RTC rows drive when gallons are
    # consumed from each supplier invoice.
    # ══════════════════════════════════════════════════════════════════════════════
    bol_alloc = {}   # bol_str → {inv_str, batch_str, cost_usd, cost_mxn}
 
    for i, row in enumerate(ws_lt_r.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not row[0]: continue
        grp = row[10]
        if grp not in ("RTB", "RTC"):
            continue
 
        bol_str  = str(row[30]).strip() if row[30] else ""
        info     = bol_info.get(bol_str)
        if not info:
            continue
 
        supplier = info["supplier"]
        bol_gals = info["gals"]
        if bol_gals <= 0:
            continue
 
        queue = supplier_queues.get(supplier, [])
        remaining = bol_gals
        alloc_usd, alloc_mxn, inv_labels, batch_labels = [], [], [], []
 
        for inv in queue:
            if remaining <= 1e-6: break
            if inv["avail"] <= 1e-6: continue
            draw = min(remaining, inv["avail"])
            alloc_usd.append((draw, inv["rate_usd_gal"]))
            alloc_mxn.append((draw, inv["supply_cost_mxn"]))
            if inv["inv_num"] not in inv_labels:
                inv_labels.append(inv["inv_num"])
            if inv["batch"] not in batch_labels:
                batch_labels.append(inv["batch"])
            inv["avail"] -= draw
            inv["drawn"] += draw
            remaining    -= draw
 
        bol_alloc[bol_str] = {
            "inv_str":   join_unique(inv_labels),
            "batch_str": join_unique(batch_labels),
            "cost_usd":  weighted_avg(alloc_usd),
            "cost_mxn":  weighted_avg(alloc_mxn),
        }
 
    # ══════════════════════════════════════════════════════════════════════════════
    # Write Stage 1 results back to Purchase to BOL-RTB (in its own row order)
    # D(4)=SupplierInvoice, J(10)=Cost/GalUSD, R(18)=Total/LiterMXN
    # ══════════════════════════════════════════════════════════════════════════════
    for i, row in enumerate(ws_bol_r.iter_rows(values_only=True), start=1):
        if i <= 7: continue
        if not row[0]: break
        bol_str = str(row[4]).strip() if row[4] else ""
        alloc   = bol_alloc.get(bol_str, {})
        if not alloc: continue
        ws_bol.cell(row=i, column=4).value  = alloc["inv_str"]
        ws_bol.cell(row=i, column=15).value = alloc["batch_str"]  # O = Batch
        ws_bol.cell(row=i, column=10).value = round(alloc["cost_usd"], 6)
        ws_bol.cell(row=i, column=18).value = round(alloc["cost_mxn"], 6)
 
    # Write Net RTB Gallons, Remainder, and Liter formulas to Supplier Invoices
    # W(23)=NetRTBGallons, X(24)=RemainderGallons, U(21)=formula, V(22)=formula
    for inv in inv_entries.values():
        rem = inv["orig"] - inv["drawn"]
        r   = inv["excel_row"]
        ws_inv.cell(row=r, column=23).value = round(inv["drawn"], 6)
        ws_inv.cell(row=r, column=24).value = round(rem, 6)
        ws_inv.cell(row=r, column=21).value = f"=W{r}*3.7854"
        ws_inv.cell(row=r, column=22).value = f"=X{r}*3.7854"
 
    # ══════════════════════════════════════════════════════════════════════════════
    # STAGE 2 — RTB → BTC FIFO on Load Tracking (sheet order = master order)
    # Load Tracking (0-indexed): K(10)=Groups, L(11)=LocationName, O(14)=LocCity,
    #   P(15)=Product, W(22)=NetLiters, Z(25)=TerminalName, AE(30)=BOL,
    #   AR(43)=SupplyCost, AV(47)=TotalCost/L
    # Write-back (1-indexed): AR(44), BE(57)=Batch, BF(58)=SupplierInv, BG(59)=BOLSource
    # ══════════════════════════════════════════════════════════════════════════════
    ws_lt.cell(row=1, column=59).value = "BOL Source"
    ws_lt.cell(row=1, column=60).value = "Batch Source"
 
    inventory = {}   # (product, bulk_plant) → deque of slots
    fifo_log  = []
 
    for i, row in enumerate(ws_lt_r.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not row[0]: continue
 
        grp           = row[10]
        product       = str(row[15]).strip() if row[15] else "Unknown"
        location_name = str(row[11]).strip() if row[11] else "Unknown"
        terminal_name = str(row[25]).strip().lstrip("* ") if row[25] else "Unknown"
        location_city = str(row[14]).strip() if row[14] else ""
        ld_num        = row[0]
        bol           = str(row[30]).strip() if row[30] else ""
        pickup        = row[3]
 
        try:
            net_liters = float(row[22]) if row[22] else 0.0
        except (TypeError, ValueError):
            net_liters = 0.0
 
        bulk_plant = location_name if grp == "RTB" else terminal_name
        key = (product, bulk_plant)
 
        if grp == "RTB":
            supply_cost  = float(row[43]) if row[43] else 0.0
            total_cost_l = float(row[47]) if row[47] else 0.0
            alloc        = bol_alloc.get(bol, {})
            batch_str    = alloc.get("batch_str", "")
            inv_str      = alloc.get("inv_str", "")
 
            if key not in inventory:
                inventory[key] = deque()
            supplier_upper = str(row[28]).strip().upper() if row[28] else ""
            inventory[key].append({
                "liters": net_liters, "cost": total_cost_l,
                "bol": bol, "batch": batch_str, "inv": inv_str,
                "supplier_upper": supplier_upper,
            })
 
            ws_lt.cell(row=i, column=44).value = supply_cost
            ws_lt.cell(row=i, column=57).value = batch_str   # BE Batch
            ws_lt.cell(row=i, column=58).value = inv_str     # BF Supplier Invoice
            ws_lt.cell(row=i, column=59).value = ""          # BG BOL Source
            ws_lt.cell(row=i, column=60).value = ""          # BH Batch Source
 
            queue_rem = sum(s["liters"] for s in inventory[key])
            fifo_log.append({
                "type": "RTB", "ld": ld_num, "pickup": pickup,
                "product": product, "bulk_plant": bulk_plant,
                "delivery_city": location_city, "bol": bol,
                "liters": net_liters, "cost_per_l": total_cost_l,
                "total_cost": net_liters * total_cost_l,
                "source_bols": "-", "queue_rem": queue_rem,
            })
 
        elif grp == "RTC":
            alloc     = bol_alloc.get(bol, {})
            ws_lt.cell(row=i, column=57).value = alloc.get("batch_str", "")  # BE Batch
            ws_lt.cell(row=i, column=58).value = alloc.get("inv_str", "")    # BF Supplier Invoice
            ws_lt.cell(row=i, column=59).value = ""                           # BG BOL Source
            ws_lt.cell(row=i, column=60).value = ""                           # BH Batch Source
 
        elif grp == "BTC":
            remaining, allocations = net_liters, []
            source_bols, source_batches = [], []
 
            q = inventory.get(key, deque())
            while remaining > 1e-6 and q:
                slot = q[0]
                draw = min(remaining, slot["liters"])
                allocations.append((draw, slot["cost"]))
                if slot["bol"] not in source_bols:
                    source_bols.append(slot["bol"])
                for b in slot["batch"].split(" | "):
                    b = b.strip()
                    if b and b not in source_batches:
                        source_batches.append(b)
                slot["liters"] -= draw
                remaining      -= draw
                if slot["liters"] <= 1e-6:
                    q.popleft()
 
            if remaining > 1e-6:
                allocations.append((remaining, 0.0))
                source_bols.append("No RTB")
 
            cost_per_l = weighted_avg(allocations)
            bols_str   = join_unique(source_bols)
            batch_str  = join_unique(source_batches)
            queue_rem  = sum(s["liters"] for s in inventory.get(key, deque()))
 
            ws_lt.cell(row=i, column=44).value = round(cost_per_l, 6)
            ws_lt.cell(row=i, column=57).value = ""           # BE Batch (blank for BTC)
            ws_lt.cell(row=i, column=58).value = ""           # BF Supplier Invoice (blank for BTC)
            ws_lt.cell(row=i, column=59).value = bols_str     # BG BOL Source
            ws_lt.cell(row=i, column=60).value = batch_str    # BH Batch Source
 
            fifo_log.append({
                "type": "BTC", "ld": ld_num, "pickup": pickup,
                "product": product, "bulk_plant": bulk_plant,
                "delivery_city": location_city, "bol": bol,
                "liters": net_liters, "cost_per_l": cost_per_l,
                "total_cost": net_liters * cost_per_l,
                "source_bols": bols_str, "queue_rem": queue_rem,
            })
 
    # ══════════════════════════════════════════════════════════════════════════════
    # FIFO sheet
    # ══════════════════════════════════════════════════════════════════════════════
    FILL_RTB    = PatternFill("solid", fgColor="C6EFCE")
    FILL_BTC    = PatternFill("solid", fgColor="FFDDC1")
    FILL_HDR    = PatternFill("solid", fgColor="2F5496")
    FILL_SUBHDR = PatternFill("solid", fgColor="BDD7EE")
    thin        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    def hdr(cell, val):
        cell.value = val; cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
 
    def sub(cell, val):
        cell.value = val; cell.fill = FILL_SUBHDR
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
 
    def dat(cell, val, num_fmt=None, fill=None):
        cell.value = val; cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        if num_fmt: cell.number_format = num_fmt
        if fill:    cell.fill = fill
 
    if "FIFO" in wb.sheetnames:
        del wb["FIFO"]
    ws_f = wb.create_sheet("FIFO")
 
    COLS = [
        ("Load #",             11), ("Date",              12), ("Type",          7),
        ("Product",            12), ("Bulk Plant",        12), ("Delivery City", 18),
        ("BOL",                12), ("Liters",            14), ("Cost / L (MXN)",15),
        ("Total Cost (MXN)",   17), ("Source BOLs",       36), ("Queue Bal. (L)",17),
    ]
 
    ws_f.merge_cells("A1:L1")
    c = ws_f["A1"]
    c.value = "VBP — FIFO Inventory Cost Allocation"
    c.fill  = FILL_HDR; c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_f.row_dimensions[1].height = 26
 
    for ci, (name, width) in enumerate(COLS, start=1):
        hdr(ws_f.cell(row=2, column=ci), name)
        ws_f.column_dimensions[get_column_letter(ci)].width = width
    ws_f.row_dimensions[2].height = 32
 
    running = {}
    for ri, entry in enumerate(fifo_log, start=3):
        rkey = (entry["product"], entry["bulk_plant"])
        fill = FILL_RTB if entry["type"] == "RTB" else FILL_BTC
        if rkey not in running: running[rkey] = 0.0
        running[rkey] += entry["liters"] if entry["type"] == "RTB" else -entry["liters"]
        for ci, (val, fmt) in enumerate([
            (entry["ld"],            None),
            (entry["pickup"],        "DD/MM/YYYY"),
            (entry["type"],          None),
            (entry["product"],       None),
            (entry["bulk_plant"],    None),
            (entry["delivery_city"], None),
            (entry["bol"],           None),
            (entry["liters"],        "#,##0.00"),
            (entry["cost_per_l"],    "#,##0.0000"),
            (entry["total_cost"],    "$#,##0.00"),
            (entry["source_bols"],   None),
            (running[rkey],          "#,##0.00"),
        ], start=1):
            dat(ws_f.cell(row=ri, column=ci), val, num_fmt=fmt, fill=fill)
        ws_f.row_dimensions[ri].height = 16
 
    ws_f.freeze_panes = "A3"
 
    sr = len(fifo_log) + 4
    ws_f.cell(row=sr, column=1).value = "INVENTORY REMAINING IN QUEUE"
    ws_f.cell(row=sr, column=1).font  = Font(bold=True, size=10)
    sr += 1
    for ci, label in enumerate(["Product", "Bulk Plant", "Liters in Queue", "Next Cost/L (MXN)", "Avg Cost in Inventory (MXN/L)"], start=1):
        sub(ws_f.cell(row=sr, column=ci), label)
    for (prod, bp), q in inventory.items():
        if not q: continue
        sr += 1
        dat(ws_f.cell(row=sr, column=1), prod,                        fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=2), bp,                          fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=3), sum(s["liters"] for s in q), num_fmt="#,##0.00",   fill=FILL_SUBHDR)
        dat(ws_f.cell(row=sr, column=4), q[0]["cost"],                num_fmt="#,##0.0000", fill=FILL_SUBHDR)
 
    # ══════════════════════════════════════════════════════════════════════════════
    # FIFO sheet — Average Cost in Remaining Inventory (col E of summary block)
    # ══════════════════════════════════════════════════════════════════════════════
    # inventory dict is still in scope: (product, bulk_plant) → deque of slots
    # The summary block starts at row (len(fifo_log) + 4); data rows begin 2 after that.
    # We need to find and fill col E for each inventory summary row.
    # Re-derive the summary start row the same way the sheet builder did.
    summary_data_start = len(fifo_log) + 6   # row 27=label, 28=header, 29+=data
 
    inv_items = [(k, q) for k, q in inventory.items() if q]
    for idx, ((prod, bp), q) in enumerate(inv_items):
        row_num = summary_data_start + idx
        total_l   = sum(s["liters"] for s in q)
        avg_cost  = sum(s["liters"] * s["cost"] for s in q) / total_l if total_l else 0.0
        ws_f.cell(row=row_num, column=5).value = round(avg_cost, 6)
        # also style it to match the other summary cells
        dat(ws_f.cell(row=row_num, column=5), round(avg_cost, 6),
            num_fmt="#,##0.0000", fill=FILL_SUBHDR)
 
    # ══════════════════════════════════════════════════════════════════════════════
    # Overall Summary — H (Remaining Gallons in Inventory) and
    #                   I (Weighted Average Cost in Inventory)
    # Aggregate remaining inventory slots by supplier (case-insensitive match).
    # Gallons = liters / 3.7854
    # ══════════════════════════════════════════════════════════════════════════════
    LITERS_PER_GAL = 3.7854
 
    # Build supplier → [(liters, cost)] from remaining inventory slots
    from collections import defaultdict as _dd
    supplier_remaining = _dd(list)   # supplier_upper → [(liters, cost_per_l)]
 
    for (prod, bp), q in inventory.items():
        for slot in q:
            sup_upper = slot.get("supplier_upper", "")
            if sup_upper:
                supplier_remaining[sup_upper].append((slot["liters"], slot["cost"]))
 
    ws_os = wb["Overall Summary"]
 
    # Find supplier rows (rows 4-7) and Grand Total (row 8)
    grand_liters, grand_cost_pairs = 0.0, []
 
    for i, row in enumerate(ws_os.iter_rows(values_only=False), start=1):
        if i < 4: continue
        label = str(row[0].value).strip() if row[0].value else ""
        if not label or label == "Row Labels":
            continue
 
        label_upper = label.upper()
 
        if label_upper == "GRAND TOTAL":
            # Write grand total after supplier loop
            grand_gals    = grand_liters / LITERS_PER_GAL
            grand_avg_cost = (sum(l * c for l, c in grand_cost_pairs) /
                              sum(l for l, _ in grand_cost_pairs)
                              if grand_cost_pairs else 0.0)
            ws_os.cell(row=i, column=8).value = round(grand_gals, 6)
            ws_os.cell(row=i, column=9).value = round(grand_avg_cost, 6)
            break
 
        # Find matching supplier in remaining inventory (case-insensitive)
        matched_slots = []
        for sup_upper, slots in supplier_remaining.items():
            if sup_upper in label_upper or label_upper in sup_upper:
                matched_slots.extend(slots)
 
        total_l = sum(l for l, _ in matched_slots)
        avg_c   = (sum(l * c for l, c in matched_slots) / total_l
                   if total_l else 0.0)
        gals    = total_l / LITERS_PER_GAL
 
        ws_os.cell(row=i, column=8).value = round(gals, 6)
        ws_os.cell(row=i, column=9).value = round(avg_c, 6) if total_l else 0.0
 
        grand_liters += total_l
        grand_cost_pairs.extend(matched_slots)
 
    wb.save(DST)
    print("Done")
 
 
 
if __name__ == "__main__":
    run_fifo(
        "/mnt/user-data/uploads/Investor_Summary_FIFO_7.xlsx",
        "/mnt/user-data/outputs/Investor_Summary_FIFO.xlsx"
    )
